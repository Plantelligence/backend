"""
Servico de exportacao de relatorios em CSV, XLSX e PDF.

Gera arquivos binarios prontos para download a partir de uma lista de
relatorios (modelos SQLAlchemy Relatorio).
"""

from __future__ import annotations

import csv
import io
from typing import Any


def _format_value(relatorio: Any, field: str) -> str:
    """Extrai e formata um campo do relatorio como string."""
    val = getattr(relatorio, field, None)
    if val is None:
        return ""
    return str(val)


def _periodo_label(relatorio: Any) -> str:
    inicio = getattr(relatorio, "periodo_inicio", "")
    fim = getattr(relatorio, "periodo_fim", "")
    return f"{inicio} ate {fim}"


def _relatorio_rows(relatorios: list) -> list[dict]:
    """Converte relatorios em lista de dicionarios para exportacao."""
    rows = []
    for r in relatorios:
        rows.append({
            "Periodo": _periodo_label(r),
            "Temperatura media (C)": _format_value(r, "avg_temperatura"),
            "Umidade do ar media (%)": _format_value(r, "avg_umidade"),
            "Umidade do solo media (%)": _format_value(r, "avg_umidade_solo"),
            "Luminosidade media (lux)": _format_value(r, "avg_luminosidade"),
            "Resumo": _format_value(r, "resumo") or "",
            "Criado em": _format_value(r, "criado_em"),
            "Auto-gerado": "Sim" if getattr(r, "auto_generated", False) else "Nao",
        })
    return rows


# ── CSV ─────────────────────────────────────────────────────────────────────

def export_csv(relatorios: list, estufa_nome: str) -> bytes:
    """Gera CSV com os relatorios da estufa."""
    rows = _relatorio_rows(relatorios)
    if not rows:
        return b""

    output = io.StringIO()
    fieldnames = list(rows[0].keys())
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)

    csv_bytes = output.getvalue().encode("utf-8-sig")
    output.close()
    return csv_bytes


# ── XLSX ────────────────────────────────────────────────────────────────────

def export_xlsx(relatorios: list, estufa_nome: str) -> bytes:
    """Gera XLSX com os relatorios da estufa usando openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    rows = _relatorio_rows(relatorios)

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorios"

    # estilos
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Arial", size=10)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # cabecalho
    if rows:
        for col_idx, header in enumerate(rows[0].keys(), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # dados
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, (key, value) in enumerate(row_data.items(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = cell_font
                cell.alignment = cell_align
                cell.border = thin_border

    # largura das colunas
    col_widths = [22, 20, 20, 22, 22, 50, 22, 14]
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + idx) if idx <= 26 else "A" + chr(64 + idx - 26)].width = width

    # buffer
    from openpyxl.utils import get_column_letter
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── PDF ─────────────────────────────────────────────────────────────────────

def export_pdf(relatorios: list, estufa_nome: str) -> bytes:
    """Gera PDF com os relatorios da estufa usando reportlab."""
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
        PageBreak,
    )
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_JUSTIFY

    buf = BytesIO()

    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2.5 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="TitleCustom",
        parent=styles["Title"],
        fontSize=20,
        textColor=colors.HexColor("#991B1B"),
        spaceAfter=4 * mm,
    ))
    styles.add(ParagraphStyle(
        name="Subtitle",
        parent=styles["Normal"],
        fontSize=11,
        textColor=colors.HexColor("#6B7280"),
        spaceAfter=8 * mm,
    ))
    styles.add(ParagraphStyle(
        name="BodyCustom",
        parent=styles["Normal"],
        fontSize=9,
        leading=13,
        alignment=TA_LEFT,
    ))

    elements = []

    # titulo
    elements.append(Paragraph("Relatorios da Estufa", styles["TitleCustom"]))
    elements.append(Paragraph(estufa_nome, styles["Subtitle"]))
    elements.append(Spacer(1, 4 * mm))

    if not relatorios:
        elements.append(Paragraph("Nenhum relatorio encontrado.", styles["BodyCustom"]))
    else:
        # cabecalho da tabela
        headers = [
            "Periodo",
            "Temp.\n(C)",
            "Umidade\n(%)",
            "Um. Solo\n(%)",
            "Luminos.\n(lux)",
            "Resumo",
            "Criado em",
        ]
        data = [headers]

        for r in relatorios:
            data.append([
                _periodo_label(r),
                _format_value(r, "avg_temperatura") or "—",
                _format_value(r, "avg_umidade") or "—",
                _format_value(r, "avg_umidade_solo") or "—",
                _format_value(r, "avg_luminosidade") or "—",
                (_format_value(r, "resumo") or "—")[:120],
                _format_value(r, "criado_em"),
            ])

        col_widths = [
            3.2 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 2.0 * cm, 6.5 * cm, 2.5 * cm,
        ]

        table = Table(
            data,
            colWidths=col_widths,
            repeatRows=1,
        )

        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#991B1B")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FFFFFF")),
            ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#1F2937")),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (0, 1), (5, -1), "LEFT"),
            ("ALIGN", (6, 1), (6, -1), "CENTER"),
            ("VALIGN", (0, 1), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#FFFFFF"), colors.HexColor("#F9FAFB")]),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))

        elements.append(table)

    doc.build(elements)
    buf.seek(0)
    return buf.getvalue()
